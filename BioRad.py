import csv
import pathlib
import sqlite3
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import tkinter as tk
import tkinter.font as tkFont
from tkinter import filedialog, messagebox
from pathlib import Path
from datetime import datetime, timedelta
import math
import random
import shutil
from itertools import islice


"""
  Plan on using Openpyxl later use Gspread for Google Sheets
  Sqlite3,CSV, TKinter, OS, Random and Path are a few libraries used to accomplish some
  pretty amazing data manipulations and presentations. 
  Will Create Several Tables for INNER JOIN statement/s
  https://dbschema.com/database-designer/Sqlite.html 15 Day Trial for database diagram
  There are some file conventions that are used to open excel and text data
  Though there is flexibility in the selection of the results selection
  Probably need to check if Directory Exists and DB exists
  IF not Make Directory and DB and Create Tables.
  Really Should be there as this would be on the production side and created during installation
  However Create if NOT EXISTS Maybe useful for a different program try: except: blocks
  Plan on using random number and Date to string to create sample Machine Results tab
  delimited files for manual loading a table and testing the program.
  
  Trying to Decide whether to copy a blank workbook as a template or
  create new workbook from Openpyxl to use for Data Entry???
  
  Tab Delimited File Has HC High Control and LC Low Control for result verification
  
  Will be using a larger Font for the workbook to assist for visual impairments Calibri 18 PT
  Many Fields can be blank as fields just aren't used by most clients or are for future use
  or no longer in use.
  At this time the Excel File will be used for at most 200 Records so A1: AA200 is the range

"""

conn = sqlite3.connect('C:\\BioRad\\BioRad.db')
cur = conn.cursor()



simulateResults = []
simulateResults.extend(range(1,500)) # First Param is optional and starts at zero.
# Range is immutable so need to shuffle so assigned range to list
random.shuffle(simulateResults)
print(simulateResults)
dateList = []

ExcelHeader = ["SampleID","DateReported","A1cResult","Client Code","eAG","Member ID","Date Collected","Height Feet",
               "Height Inches","Inches Total","Weight","BMI","HemoglobinA1cComment","First Name","Last Name",
               "Middle Initial","Street Number","Apt Num","City","State","Zip","Fax","Sex","Email","A1cReportOnly",
               "Practitioner","Albumin"]

SampleTabLine1 = "CDM Export A1c Laboratories Instrument #1 07/23/2021 Run #25 V2TURBO_A1c\n"


SampleTabLine3 = "Run #	Inj #	RackID	Type	Sample ID / Lot Number	Injection Date	Injection Time	Peak Name	RT	" \
                 "Concentration % NGSP\n"
print(SampleTabLine3)
SampleTabFiller1 = """25	1025	0001	LC	85821	07/23/2021	11:49:26	A1a	0.156	
25	1025	0001	LC	85821	07/23/2021	11:49:26	A1b	0.202	
25	1025	0001	LC	85821	07/23/2021	11:49:26	F	0.251	
25	1025	0001	LC	85821	07/23/2021	11:49:26	LA1c	0.358	
25	1025	0001	LC	85821	07/23/2021	11:49:26	A1c	0.435	5.4 
25	1025	0001	LC	85821	07/23/2021	11:49:26	P3	0.747	
25	1025	0001	LC	85821	07/23/2021	11:49:26	P4	0.813	
25	1025	0001	LC	85821	07/23/2021	11:49:26	Unknown	0.963	
25	1026	0001	HC	85822	07/23/2021	11:51:02	A1a	0.157	
25	1026	0001	HC	85822	07/23/2021	11:51:02	A1b	0.206	
25	1026	0001	HC	85822	07/23/2021	11:51:02	F	0.255	
25	1026	0001	HC	85822	07/23/2021	11:51:02	LA1c	0.372	
25	1026	0001	HC	85822	07/23/2021	11:51:02	A1c	0.448	10.1 
25	1026	0001	HC	85822	07/23/2021	11:51:02	P3	0.759	
25	1026	0001	HC	85822	07/23/2021	11:51:02	P4	0.831	
25	1026	0001	HC	85822	07/23/2021	11:51:02	Ao	0.975
"""
print(SampleTabFiller1)
SampleTabFiller2 = """25	1027	0001	P	461a	07/23/2021	11:52:38	A1a	0.157	
25	1027	0001	P	461a	07/23/2021	11:52:38	A1b	0.207	
25	1027	0001	P	461a	07/23/2021	11:52:38	LA1c	0.365	
"""
print(SampleTabFiller2)

SampleTabFiller3 = """25	1027	0001	P	461a	07/23/2021	11:54:14	P3	0.767	
25	1027	0001	P	461a	07/23/2021	11:54:14	P4	0.832	
25	1027	0001	P	461a	07/23/2021	11:54:14	Ao	0.972	
"""
print(SampleTabFiller3)



"""
Sample Tab Delimited File Looks like this
25	1025	0001	LC	85821	07/23/2021	11:49:26	A1a	0.156	
25	1025	0001	LC	85821	07/23/2021	11:49:26	A1b	0.202	
25	1025	0001	LC	85821	07/23/2021	11:49:26	F	0.251	
25	1025	0001	LC	85821	07/23/2021	11:49:26	LA1c	0.358	
25	1025	0001	LC	85821	07/23/2021	11:49:26	A1c	0.435	5.4 
25	1025	0001	LC	85821	07/23/2021	11:49:26	P3	0.747	
25	1025	0001	LC	85821	07/23/2021	11:49:26	P4	0.813	
25	1025	0001	LC	85821	07/23/2021	11:49:26	Unknown	0.963	
25	1026	0001	HC	85822	07/23/2021	11:51:02	A1a	0.157	
25	1026	0001	HC	85822	07/23/2021	11:51:02	A1b	0.206	
25	1026	0001	HC	85822	07/23/2021	11:51:02	F	0.255	
25	1026	0001	HC	85822	07/23/2021	11:51:02	LA1c	0.372	
25	1026	0001	HC	85822	07/23/2021	11:51:02	A1c	0.448	10.1 
25	1026	0001	HC	85822	07/23/2021	11:51:02	P3	0.759	
25	1026	0001	HC	85822	07/23/2021	11:51:02	P4	0.831	
25	1026	0001	HC	85822	07/23/2021	11:51:02	Ao	0.975	
25	1027	0001	P	461	07/23/2021	11:52:38	A1a	0.157	
25	1027	0001	P	461	07/23/2021	11:52:38	A1b	0.207	
25	1027	0001	P	461	07/23/2021	11:52:38	LA1c	0.365	
25	1027	0001	P	461	07/23/2021	11:52:38	A1c	0.450	9.0
25	1027	0001	P	461	07/23/2021	11:52:38	P3	0.751	
25	1027	0001	P	461	07/23/2021	11:52:38	P4	0.833	
25	1027	0001	P	461	07/23/2021	11:52:38	Ao	0.978	
25	1028	0001	P	191034	07/23/2021	11:54:14	Unknown	0.111	
25	1028	0001	P	191034	07/23/2021	11:54:14	A1a	0.152	
25	1028	0001	P	191034	07/23/2021	11:54:14	A1b	0.208	
25	1028	0001	P	191034	07/23/2021	11:54:14	F	0.258	
25	1028	0001	P	191034	07/23/2021	11:54:14	LA1c	0.378	
25	1028	0001	P	191034	07/23/2021	11:54:14	A1c	0.452	8.5
25	1028	0001	P	191034	07/23/2021	11:54:14	P3	0.767	
25	1028	0001	P	191034	07/23/2021	11:54:14	P4	0.832	
25	1028	0001	P	191034	07/23/2021	11:54:14	Ao	0.972	
"""



def roundUP(x):
    return int(math.ceil(x / 100.0)) * 100
    # https://stackoverflow.com/questions/8866046/python-round-up-integer-to-next-hundred

def dataSimulation():
    # Create Folder for tab delimited text files.
    # Change CWD to C:\BioRad Folder
    # Will Need to use multiple strategies to try to memic real world scenarios
    path = pathlib.Path("C:\BioRad\SampleData")
    # Chack Path exists if not create
    path.mkdir(parents=True, exist_ok=True)
    print("Help Button Pressed.  ")
    # Create orders for active Clients Round up to nearest 10
    # Not all samples will be returned
    # Create a Random Number b/t 75 and 200 for batch processing
    # There will not be more than 200 samples processed per batch
    # From that number divide by the active count of clients
    # Will be looping a set number of days and each iteration will be for one day data population
    #
    # Testing Modified Date by one day for use in file creation.
    dDate = datetime.now()
    dDate = dDate.strftime("%m/%d/%Y")
    dateList.append(dDate)
    # print(dDate) Date is represented at m/d/Y 09/06/2021

    for i in range(7):
        date = datetime.strptime(dDate, "%m/%d/%Y")
        modified_date = date - timedelta(days=1)
        dDate = datetime.strftime(modified_date, "%m/%d/%Y")
        dateList.append(dDate)
        print(datetime.strftime(modified_date, "%m/%d/%Y"))

    # Need to Get Active Count on Clients
    # Need to create variable to hold tabbed data before result and after result for sample num
    print(len(dateList))
    random_number = 0
    for list in dateList:
        # Create File for date name results_m/d/Y
        print(list)
        random_number = random.randint(65, 200)
        cur.execute("SELECT ID,ClientCode FROM Clients WHERE Active = 1")
        data = cur.fetchall()
        clientDictionary = {}
        for row in data:
            clientDictionary[row[0]] = str(row[1])
        print(clientDictionary)
        print(len(data))
        lenData = len(data)
        samples = 0
        samples = round(random_number / len(data))
        print("Total Number of round(random_number / len(data)): " + str(samples))


        # multiply lenData * samples
        val1 = (lenData * samples) + lenData
        val1 = round(val1/10)*10
        print(str(val1))
        print("SamplesNum " + str(val1))
        cur.execute("SELECT SampleNum FROM SampleSeed")
        datarecord = cur.fetchall() # possibly use fetchone here?
        for row in datarecord:
            print(str(row[0]))
            dataS = row[0]
            # update SampleNum to new value
            dataSE = row[0] + val1
        cur.execute('''UPDATE SampleSeed SET 
                          SampleNum = ?
                        WHERE SampleNum = ?''', (dataSE, dataS))
        conn.commit()


        #if len(data) == 0:



        # From this number divide by active clients then round up to nearest whole number
        # Assign to master order file
        # Write to file
        random_number = random_number - lenData
        print(random_number)
        sampleStartNum = dataS
        sampleEndNum = sampleStartNum + samples

        for key in clientDictionary:
            cur.execute("INSERT INTO MasterOrder (ClientID, ClientCode, SampleStartNum, SampleEndNum) VALUES(?, ?, ?, ?)", (key, clientDictionary[key], sampleStartNum, sampleEndNum))
            conn.commit()
            sampleStartNum = sampleEndNum + 1
            sampleEndNum = sampleStartNum + samples



        # Now create the text files based on dates
        # Create a range from 1 to random_number then shuffle in a list
        sampleResults = []
        print(dataS)

        sampleResults.extend(range(dataS, (random_number+dataS),1))  # First Param is optional and starts at zero.
        # Range is immutable so need to shuffle so assigned range to list
        random.shuffle(sampleResults)
        print(sampleResults)
        # Let go ahead and write orders to MasterOrder based on samples integer

        random_a1c = 0.0
        random_a1c = round(random.uniform(4.2, 10.5), 1)  # Gets number to one decimal precision using round params
        print("A1c: " + str(random_a1c))
        # This is used to create a A1c value.

        dateName = str(list).replace("/","_")
        print(dateName)
        file = open("C:\\BioRad\\SampleData\\bioradRun_" + dateName + ".txt", 'w+')

        SampleTabLineN1 = SampleTabLine1.replace("07/23/2021",str(list))
        file.write(SampleTabLineN1)
        file.write("\n")
        file.write(SampleTabLine3)
        file.write("\n")
        SampleTabFillerN1 = SampleTabFiller1.replace("07/23/2021", str(list))
        file.write(SampleTabFillerN1)

        for list2 in sampleResults:
            SampleTabFillerN2 = SampleTabFiller2.replace("07/23/2021", str(list))
            SampleTabFillerN2 = SampleTabFiller2.replace("461a", str(list2))
            file.write(SampleTabFillerN2)
            # "25	1025	0001	LC	85821	07/23/2021	11:49:26	A1c	0.435	5.4*"
            # With or without asterisks
            lineData = "25\t1027\t0001\tP\t" + str(list2) + "\t" + list + "\t11:49:26\tA1c\t0.435\t" + str(random_a1c) + chr(42) + "\n"
            #print(lineData)
            file.write(lineData)
            SampleTabFillerN3 = SampleTabFiller3.replace("07/23/2021", str(list))
            SampleTabFillerN3 = SampleTabFiller3.replace("461a", str(list2))
            file.write(SampleTabFillerN3)
            random_a1c = round(random.uniform(3.2, 10.5), 1)


        file.close()





# Problem Putting in DB is DB doesn't exist or was deleted then would have to recreate the database with the app.
#



class App:
    def __init__(self, root):
        # setting title
        root.title("A1c Batch Data")
        # setting window size
        width=303
        height=254
        screenwidth = root.winfo_screenwidth()
        screenheight = root.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        root.geometry(alignstr)
        root.resizable(width=False, height=False)
        excel_file_path = "C:\\BioRad\\Book1.xlsx"
        switchExcel = 0
        if os.path.isfile(excel_file_path):
            switchExcel = 1
            # Button Overlay

        csv_file_path = "C:\\BioRad\\BioRad.csv"
        greyTopBtn = 0
        if os.path.isfile(csv_file_path):
            greyTopBtn = 1





        self.btnOpenExcel1 = tk.Button(root)
        self.btnOpenExcel1["bg"] = "#009688"
        ft = tkFont.Font(family='Times', size=10)
        self.btnOpenExcel1["font"] = ft
        self.btnOpenExcel1["fg"] = "#000000"
        self.btnOpenExcel1["justify"] = "center"
        self.btnOpenExcel1["text"] = "Open Book1 for Input"
        self.btnOpenExcel1.place(x=20, y=90, width=260, height=40)
        self.btnOpenExcel1["command"] = self.btnOpenExcel1_command

        self.btnOpenExcel = tk.Button(root)
        self.btnOpenExcel["bg"] = "#009688"
        ft = tkFont.Font(family='Times', size=10)
        self.btnOpenExcel["font"] = ft
        self.btnOpenExcel["fg"] = "#000000"
        self.btnOpenExcel["justify"] = "center"
        self.btnOpenExcel["text"] = "Open Excel for Input"
        self.btnOpenExcel.place(x=20, y=90, width=260, height=40)
        self.btnOpenExcel["command"] = self.btnOpenExcel_command

        self.btnProcessExcel=tk.Button(root)
        #btnProcessExcel["state"] = "disable"
        self.btnProcessExcel["bg"] = "#009688"
        ft = tkFont.Font(family='Times',size=10)
        self.btnProcessExcel["font"] = ft
        self.btnProcessExcel["fg"] = "#000000"
        self.btnProcessExcel["justify"] = "center"
        self.btnProcessExcel["text"] = "Process Excel"
        self.btnProcessExcel.place(x=20,y=150,width=260,height=40)
        self.btnProcessExcel["command"] = self.btnProcessExcel_command



        self.btnBioRad = tk.Button(root)
        self.btnBioRad["activebackground"] = "#009688"
        self.btnBioRad["activeforeground"] = "#2e3445"
        self.btnBioRad["bg"] = "#009688"
        ft = tkFont.Font(family='Times',size=10)
        self.btnBioRad["font"] = ft
        self.btnBioRad["fg"] = "#000000"
        self.btnBioRad["justify"] = "center"
        self.btnBioRad["text"] = "Read BioRad"
        self.btnBioRad.place(x=20,y=30,width=260,height=40)
        self.btnBioRad["command"] = self.btnBioRad_command
        if greyTopBtn == 1:
            self.btnBioRad["state"] = "disable"
            # self.btnProcessExcel["state"] = "normal"
            if switchExcel == 1:
                #self.btnOpenExcel["state"] = "normal"
                self.btnMiddle2()
            else:
                self.btnProcessExcel["state"] = "disable"
        else:
            self.btnBioRad["state"] = "normal"
            self.btnProcessExcel["state"] = "disable"
            self.btnOpenExcel["state"] = "disable"
            self.btnOpenExcel1["state"] = "disable"











        self.btnSimulate=tk.Button(root)
        config_file = "C:\\BioRad\\Config\\Config.txt"
        path = Path(config_file)
        if path.is_file():
            x = 0
            print(f'The file {config_file} exists')
            with open(config_file) as f:
                firstline = f.readline().rstrip()
                xList = firstline.split(":")
                x = int(xList[1])
                print(x)
            if x == 1:
                self.btnSimulate["state"] = "normal"
            else:
                self.btnSimulate["state"] = "disable"
        self.btnSimulate["bg"] = "#87a987"
        ft = tkFont.Font(family='Times',size=10)
        self.btnSimulate["font"] = ft
        self.btnSimulate["fg"] = "#000000"
        self.btnSimulate["justify"] = "center"
        self.btnSimulate["text"] = "Siumlate BioRad"
        self.btnSimulate.place(x=130,y=210,width=150,height=26)
        self.btnSimulate["command"] = self.btnSimulate_command
        

        self.btnHelp = tk.Button(root)
        self.btnHelp["bg"] = "#87a987"
        ft = tkFont.Font(family='Times', size=10)
        self.btnHelp["font"] = ft
        self.btnHelp["fg"] = "#000000"
        self.btnHelp["justify"] = "center"
        self.btnHelp["text"] = "Help"
        self.btnHelp.place(x=20, y=210, width=80, height=26)
        self.btnHelp["command"] = self.btnHelp_command
        filename = ""
        folder_selected = ""


    def btnBioRad_command(self):
        print("Read Biorad command")
        self.filename = filedialog.askopenfile()
        # Get Folder Path to change CWD Current Working Directory
        self.folder_selected = Path(self.filename.name).parent
        # Gets User Interaction to select the BidRad File Name and Path
        try:
            print("Select File is:", self.filename.name)
            self.readTabDelFile()
        except:
            print("No File Selected.")

    def btnProcessExcel_command(self):
        print("Process Excel command")
        self.btnOpenExcel["state"] = "disable"
        self.btnOpenExcel1["state"] = "disable"
        # From Book1.xlsx copy cells to csv file
        # Move Book1.xlsx file to Archive Directory and rename
        excel_file_path = "C:\\BioRad\\Book1.xlsx"
        path = Path(excel_file_path)
        if path.is_file():
            print(f'The file {excel_file_path} exists')
        #else:
        try:
            # Open WorkBook to see if exists
            file_name = "C:\\BioRad\\Book1.xlsx"
            wb = load_workbook(file_name, data_only=True)
            # Don't want to use this create but will copy Template
            # If Book1.xlsx Exists Grey out Button for processing.
            # Maybe a Hot Key to Reopen Book1.xlsx instead of process if it is even possible with TKinter and python?
            sheet = wb['Data']

            tempCSV = []
            for i, row in enumerate(sheet.iter_rows()):
                tempList = []
                if i == 0:
                    continue
                check = 0
                for cell in row:
                    #sheet['A1'].value
                    CellRef = "A" + str(i+1)
                    CellRef1 = "AA" + str(i + 1)
                    check = 0
                    if sheet[CellRef].value:
                        #sheet[str(cell.coordinate)].font = Font(size=18)
                        #tempList.append(cell.value)
                        if sheet[cell.coordinate].value is None:
                            tempList.append("")
                        else:
                            # Convert Date String to date
                            # datetime_object = datetime.strptime('Jun 1 2005  1:33PM', '%b %d %Y %I:%M%p')
                            # timestamp = str(now.strftime("%Y%m%d_%H-%M-%S"))
                            # https://stackoverflow.com/questions/466345/converting-string-into-datetime
                            # G Cells are dates
                            print(cell.coordinate)
                            print("i " + str(i+1))
                            print(cell.column)
                            if cell.column == 7:
                                print("format dates")
                                datev = datetime.strftime(sheet[cell.coordinate].value, '%m/%d/%Y')
                                print(datev)
                                tempList.append(datev)
                            elif cell.column == 12:
                                # https://pythonguides.com/python-print-2-decimal-places/
                                val = sheet[cell.coordinate].value
                                print(str(sheet[cell.coordinate].value))
                                # float = 2.154327
                                if sheet[cell.coordinate].value == '0.00':
                                    tempList.append("")
                                else:
                                    format_float = "{:.2f}".format(val)
                                    tempList.append(format_float)
                            elif cell.column == 10:
                                print(sheet[cell.coordinate].value)
                                if sheet[cell.coordinate].value == 0:
                                    tempList.append("")
                                else:
                                    tempList.append(sheet[cell.coordinate].value)
                            else:
                                tempList.append(sheet[cell.coordinate].value)


                        #print(cell.coordinate)
                        #print(sheet[cell.coordinate].value)
                        check = 1

                if check == 1:
                    temp_string = ""
                    # https://stackoverflow.com/questions/44778/how-would-you-make-a-comma-separated-string-from-a-list-of-strings
                    temp_string = ",".join(map(str,tempList))
                    tempCSV.append(temp_string)

            print(str(tempCSV))
            print(len(tempCSV))
            results_file_path = "C:\\BioRad\\Results\\Results.csv"
            path = Path(results_file_path)
            f = open(results_file_path, "w")
            counter = 0
            for list in tempCSV:
                counter +=1
                f.write(str(list) + "\n")
                print(str(counter))
            f.close()
            wb.save(file_name)
            # Let's Cleanup
            self.masterCleanup()

        except:
            print("Unable to Process...")

    def btnOpenExcel1_command(self):
        print("Open Excel Book1")
        file_name = "C:\\BioRad\\Book1.xlsx"
        os.system("start EXCEL.EXE " + file_name)



    def btnSimulate_command(self):
        # open WordPad to read Help Files .rtf
        print("Simulate Bio Rad Results")
        dataSimulation()



    def btnMiddle2(self):
        self.btnOpenExcel1 = tk.Button(root)
        self.btnOpenExcel1["bg"] = "#009688"
        ft = tkFont.Font(family='Times', size=10)
        self.btnOpenExcel1["font"] = ft
        self.btnOpenExcel1["fg"] = "#000000"
        self.btnOpenExcel1["justify"] = "center"
        self.btnOpenExcel1["text"] = "Open Book1 for Input"
        self.btnOpenExcel1.place(x=20, y=90, width=260, height=40)
        self.btnOpenExcel1["command"] = self.btnOpenExcel1_command
        self.btnOpenExcel1["state"] = "normal"


    def btnMiddle1(self):
        self.btnOpenExcel = tk.Button(root)
        self.btnOpenExcel["bg"] = "#009688"
        ft = tkFont.Font(family='Times', size=10)
        self.btnOpenExcel["font"] = ft
        self.btnOpenExcel["fg"] = "#000000"
        self.btnOpenExcel["justify"] = "center"
        self.btnOpenExcel["text"] = "Open Excel for Input"
        self.btnOpenExcel.place(x=20, y=90, width=260, height=40)
        self.btnOpenExcel["command"] = self.btnOpenExcel_command
        self.btnOpenExcel["state"] = "normal"

    def btnHelp_command(self):
        # open WordPad to read Help Files .rtf
        print("Help command")
        # also a simulate bio rad config file?
        help_file_name = "C:\\BioRad\\Help\\HelpFile.rtf"
        os.system("start WordPad.EXE " + help_file_name)
        #self.btnMiddle2()
        #self.masterReset()

    def ftpInput(self):
        print("Ftp Input File")
        now = datetime.now()
        timestamp = str(now.strftime("%Y%m%d_%H-%M-%S"))
        print(timestamp)
        ftpOld = "C:\\BioRad\\Results\\Results.csv"
        ftpNew = "C:\\BioRad\\FTP\\Results" + timestamp + ".csv"
        shutil.move(ftpOld, ftpNew)
        # Really would probably FTP file to FTP Server for later processing.
        # But the move is used to simulate last step here.
        # Close out of tkinter
        root.destroy()

    def masterCleanup(self):
        print("Cleanup Files")
        now = datetime.now()
        timestamp = str(now.strftime("%Y%m%d_%H-%M-%S"))
        print(timestamp)
        csvOld = "C:\\BioRad\\BioRad.csv"
        csvNew = "C:\\BioRad\\Archive\\BioRad" + timestamp + ".csv"
        xlOld = "C:\\BioRad\\Book1.xlsx"
        xlNew = "C:\\BioRad\\Archive\\Book" + timestamp + ".xlsx"
        shutil.move(csvOld, csvNew)
        shutil.move(xlOld, xlNew)
        self.ftpInput()

    def masterReset(self):
        print("To Do master reset? ")


    def btnOpenExcel_command(self):
        print("Open Excel command")

        #Check to see if Excel File is Present
        # Create a blank from copying template
        # After File has been processed move to archive directory
        excel_file_path = "C:\\BioRad\\Book1.xlsx"
        path = Path(excel_file_path)
        if path.is_file():
            print(f'The file {excel_file_path} exists')
        else:
            wb = load_workbook('C:\\BioRad\\XLTemplate\\Template.xlsx')
            wb.save('C:\\BioRad\\Book1.xlsx')

        try:
            # Open WorkBook to see if exists
            file_name = "C:\\BioRad\\Book1.xlsx"
            wb = load_workbook(file_name, data_only=True)
            # Don't want to use this create but will copy Template
            # If Book1.xlsx Exists Grey out Button for processing.
            # Maybe a Hot Key to Reopen Book1.xlsx instead of process if it is even possible with TKinter and python?
            sheet = wb['Data']
            sheet3 = wb['Chart']
            font1 = Font(size=18, color='00000000')
            column = sheet.column_dimensions['A']
            column.font = font1

            # iter_row skip first line
            # https://stackoverflow.com/questions/54261748/skip-first-row-in-openpxyl/54262175
            for i, row in enumerate(sheet.iter_rows()):
                if i == 0:
                    continue
                for cell in row:
                    sheet[str(cell.coordinate)].font = Font(size=18)
                    #sheet[str(cell.coordinate)].value = "33"

            # CWD
            csvfilename = "C:\\BioRad\\BioRad.csv"
            Normal = 0
            Prediabetes = 0
            Diabetes =0

            with open(csvfilename, newline='') as a1cxl:
                csvfile = csv.reader(a1cxl, delimiter=',')
                count = 0

                # print("made it here")
                for line in csvfile:
                    count = count + 1
                    # Get Line Count for CSV File
                    # print("count is :" + str(count))
                a1cxl.seek(0)
                rowNum = 2 # start at row 2
                for line in csvfile:
                    sheet.cell(row=rowNum, column=1).value = line[0]
                    sheet.cell(row=rowNum, column=2).value = line[1]
                    sheet.cell(row=rowNum, column=3).value = line[2]
                    if float(line[2]) <= 5.6:
                        Normal = Normal + 1
                    if 5.7 <= float(line[2]) <= 6.4:
                        Prediabetes = Prediabetes + 1
                    if float(line[2]) >= 6.5:
                        Diabetes = Diabetes + 1
                    #Create Query to search MasterOrder Table for oder between the sample number.
                    #SampleStartNum SampleEndNum
                    # SELECT ClientCode FROM MasterOrder WHERE ? BETWEEN SampleStartNum and SampleEndNum
                    # https://stackoverflow.com/questions/54261748/skip-first-row-in-openpxyl/54262175
                    #"=SUM(1, 1)"
                    sheet.cell(row=rowNum, column=5).value = "=CONCATENATE(ROUND((28.7 * C" + str(rowNum) + ")-46.7,0), \" mg/dl\")"
                    sheet.cell(row=rowNum, column=10).value = "=(H" + str(rowNum) + "* 12) + I" + str(rowNum)
                    sheet.cell(row=rowNum, column=12).value = "=IFERROR(((K" + str(rowNum) + "*1)/((J" + str(rowNum) + \
                                                              "*1)*(J" + str(rowNum) + "*1))*703),\"0.00\")"

                    try:
                        cur.execute("SELECT ClientCode FROM MasterOrder WHERE ? BETWEEN SampleStartNum and SampleEndNum", (line[0],))
                        data = cur.fetchone()
                        if len(data) == 0:
                            # Skip
                            print("Skip")
                        else:
                            print(data)
                            sheet.cell(row=rowNum, column=4).value = data[0]
                        # Can calculate eAG here

                    except:
                        print("No records in MasterOrder Table")
                    rowNum = rowNum + 1

            sheet3.cell(row=2, column=3).value = Normal
            sheet3.cell(row=3, column=3).value = Prediabetes
            sheet3.cell(row=4, column=3).value = Diabetes

            wb.save(file_name)
            os.system("start EXCEL.EXE " + file_name)
            self.btnProcessExcel["state"] = "normal"
        except:
            messagebox.showinfo("showinfo", "No Excel File Created \nFirst Read BioRad File.")

    def readTabDelFile(self):
        print("Parse BioRad File.")

        print(self.filename.name)
        listFile = []
        listSplit = []
        with open(self.filename.name, newline='') as a1cresults:
            results = csv.reader(a1cresults, delimiter='\t')
            count = 0
            for result in results:
                # print(result) Skip first 3 lines
                if count > 3:
                    if result[3] == "P":
                        if result[7] == "A1c":
                            print(result[4] + " " + result[5] + " " + result[9])
                            a1c = result[9]
                            if a1c.find("*") > 0:
                                a1c = a1c.replace("*", "")
                            else:
                                a1c = result[9]
                            listSplit.append(result[4] + "," + result[5] + "," + a1c)
                            # Check to see if record exists
                            cur.execute("SELECT SampleID FROM MachineResults WHERE SampleID = ?", (result[4],))
                            data = cur.fetchall()
                            if len(data) == 0:
                                print('Insert Record')
                                cur.execute("INSERT INTO MachineResults (SampleID, Date, A1c) VALUES(?, ?, ?)",(result[4], result[5], a1c))
                                conn.commit()
                            else:
                                print('Update Record')
                                #There is very little chance there will ever be a duplicate
                                #But a sample in theory can be run at most twice on the machine
                                #So an update is added probably never needed based on interview
                                #with Lab Technician.
                                cur.execute('''UPDATE MachineResults SET 
                                                A1c = ?, 
                                                UpdateDate = ?, 
                                                Updated = ? 
                                            WHERE SampleID = ?''', (a1c, result[5], 1, result[4]))
                                conn.commit()

                    listFile.append(results)
                    # print(type(result))
                    # print(type(results))
                # May never need a counter but just in case - counter is increments below
                count = count + 1
            # cur.execute("INSERT INTO MachineResults (SampleID, Date, A1c) VALUES ('1','2','3')")
            # conn.commit()

            #cur.close()
            #conn.close()

            # print(len(result))
            # print(len(listFile))
            # print(len(listSplit))
            # print(listSplit[0])
            # Use ListSplit to create CSV File.
            # This is CWD not from directory of file
            # Implemantation can be as follows using Path(__file__).parent already used but
            # showing again for to reiterate.
            print("File Name Only:", Path(self.filename.name).name)
            print("File Folder Path:", Path(self.filename.name).parent)
            print("Folder Path Again: " + str(self.folder_selected))
            # print(self.filename)

            print(len(listSplit))
            with open('C:\\BioRad\\BioRad.csv', 'w') as file_object:
                for list in listSplit:
                    file_object.write(str(list) + "\n")
                file_object.close()
        self.btnBioRad["state"] = "disable"
        self.btnOpenExcel["state"] = "normal"

    #def CreateSampleBioRadDate(self):
        #random.uniform(1.5, 9.5)



if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
    # root.withdraw()  # Hides the root window
    # root.wm_iconbitmap('py.ico')


